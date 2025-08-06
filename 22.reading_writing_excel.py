import openpyxl  # Import the library to read/write Excel files

book = openpyxl.load_workbook("C:/Users/BIRVA/Downloads/download.xlsx")
sheet = book.active # Gets the active sheet

Dict = {}
cell = sheet.cell(row=1, column=4)
print(cell.value)

# sheet.cell(row=2, column=2).value = "Rahul"
#
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

# print(sheet.max_column)

''' Loop to print all cell values
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)
        
'''


# Print values of row where column 1 has "Testcase2"
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)


# Create a dictionary of values from the row where column 1 is "Testcase2"
Dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2, sheet.max_column + 1): # to get columns'
            print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            print(Dict)


print(Dict)

'''

openpyxl helps in data-driven testing by reading test data from Excel files.
sheet.cell(row, column).value is the core method to access Excel cells.
You can filter rows dynamically using if conditions (e.g., by Test Case Name).
Convert test row into dictionary format to easily use it in frameworks like PyTest.
Don’t forget to call book.save() if you make changes to Excel.

'''


