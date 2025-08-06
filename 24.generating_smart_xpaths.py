
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

# book = openpyxl.load_workbook("C:/Users/BIRVA/Downloads/download.xlsx")
# sheet = book.active
# Dict = {}
#
#
file_path = "C:/Users/BIRVA/Downloads/download.xlsx"
fruit_name = "Apple"
#
# for i in range(1, sheet.max_row + 1):
#     if sheet.cell(row=1, column=i).value == "price":
#         Dict["col"] = i
#
# for i in range(1, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         if sheet.cell(row=i, column=j).value == "Apple":
#             Dict["row"] == i


driver = webdriver.Chrome()
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()

# edit the excel with updated value

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
'''
The input tag has attribute: type="file" ✅
If not present, Selenium cannot handle the native OS file window.
'''

file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")

wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)   #unpacks the tuple

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)

'''

Dynamic XPath Creation (for fruit price):

Start at the element with fruit name:
//div[text()='Apple']
Go to parent:
parent::div
Go to grandparent:
parent::div → this is the row
Navigate to sibling column where price is shown.


Use dynamic XPath to adapt to UI changes (e.g. column shift).
Upload/Download tests often combine Selenium waits, file handling, and smart element location 

'''






